#!/usr/bin/env python3
"""
Round 102 — one-time reader for seed-atlas-channel-plans.js.

Same rationale as xlsx_gen.py/qr_gen.py: this sandbox's npm registry is
blocked (no exceljs/xlsx/node-xlsx available to `require()`), but openpyxl
is already on disk in this environment's Python install. This script does
the actual .xlsx parsing; the Node seed script spawns it once, reads back
plain JSON, and does every DB/business-logic decision itself in JS (matching
this backend's existing seed-script style) — Python's job here is limited to
"turn a workbook into JSON", nothing else.

Usage: python3 xlsx_read.py <path-to-xlsx> <sheetName> [<sheetName> ...]
Writes to stdout: { "<sheetName>": { "headers": [...], "rows": [[...], ...] }, ... }
Dates become ISO 'YYYY-MM-DD' strings; everything else passes through as-is
(numbers, strings, None -> null).
"""
import sys
import json
import datetime


def cell_to_json(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime('%Y-%m-%d')
    return v


def main():
    if len(sys.argv) < 3:
        sys.stderr.write('usage: xlsx_read.py <path> <sheetName> [...]\n')
        sys.exit(1)
    xlsx_path = sys.argv[1]
    sheet_names = sys.argv[2:]

    try:
        import openpyxl
    except ImportError as e:
        sys.stderr.write(f'missing dependency: {e}\n')
        sys.exit(2)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    out = {}
    for name in sheet_names:
        if name not in wb.sheetnames:
            sys.stderr.write(f'sheet not found: {name}\n')
            sys.exit(3)
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        headers = [cell_to_json(v) for v in rows[0]] if rows else []
        data_rows = [[cell_to_json(v) for v in row] for row in rows[1:]]
        out[name] = {'headers': headers, 'rows': data_rows}

    sys.stdout.write(json.dumps(out))


if __name__ == '__main__':
    main()
